import csv
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import customtkinter as ctk

from database import get_connection


class ReportPage(ctk.CTkFrame):
    def __init__(self, master):
        super().__init__(master, fg_color="transparent")

        self.summary_var = tk.StringVar(value="Select a report to view its details.")
        self.current_report_title = None
        self.current_report_headers = ()
        self.current_report_rows = []

        self.report_definitions = [
            ("1. Sales Report", self.load_sales_report),
            ("2. Inventory Report", self.load_inventory_report),
            ("3. Product Report", self.load_product_report),
            ("4. Payment Report", self.load_payment_report),
            ("5. Customer Report", self.load_customer_report),
            ("6. Supplier Report", self.load_supplier_report),
        ]

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self._build_header()
        self._build_report_viewer()

    def _build_header(self):
        header = ctk.CTkFrame(
            self,
            fg_color="#1F2937",
            corner_radius=20,
            border_width=1,
            border_color="#334155"
        )
        header.grid(row=0, column=0, padx=24, pady=(24, 18), sticky="ew")
        header.grid_columnconfigure(0, weight=1)
        header.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(
            header,
            text="Reports",
            font=("Segoe UI", 30, "bold"),
            text_color="white"
        ).grid(row=0, column=0, padx=24, pady=(22, 6), sticky="w")

        ctk.CTkLabel(
            header,
            textvariable=self.summary_var,
            font=("Segoe UI", 14),
            text_color="#CBD5E1"
        ).grid(row=1, column=0, padx=24, pady=(0, 20), sticky="w")

        button_frame = ctk.CTkFrame(header, fg_color="transparent")
        button_frame.grid(row=2, column=0, columnspan=2, padx=20, pady=(0, 20), sticky="ew")

        for index, (label, command) in enumerate(self.report_definitions):
            ctk.CTkButton(
                button_frame,
                text=label,
                width=170,
                height=40,
                corner_radius=10,
                command=command
            ).grid(row=index // 3, column=index % 3, padx=10, pady=10, sticky="ew")

        for column in range(3):
            button_frame.grid_columnconfigure(column, weight=1)

        export_frame = ctk.CTkFrame(header, fg_color="transparent")
        export_frame.grid(row=3, column=0, columnspan=2, padx=20, pady=(0, 20), sticky="ew")
        export_frame.grid_columnconfigure(0, weight=1)
        export_frame.grid_columnconfigure(1, weight=1)
        ctk.CTkButton(export_frame, text="Export CSV", height=38, corner_radius=10, fg_color="#334155", hover_color="#475569", command=self.export_csv).grid(row=0, column=0, padx=(0, 6), sticky="e")
        ctk.CTkButton(export_frame, text="Export Excel", height=38, corner_radius=10, fg_color="#0F766E", hover_color="#115E59", command=self.export_excel).grid(row=0, column=1, padx=(6, 0), sticky="w")

    def _build_report_viewer(self):
        viewer_card = ctk.CTkFrame(
            self,
            fg_color="#1F2937",
            corner_radius=20,
            border_width=1,
            border_color="#334155"
        )
        viewer_card.grid(row=1, column=0, padx=24, pady=(0, 24), sticky="nsew")
        viewer_card.grid_columnconfigure(0, weight=1)
        viewer_card.grid_rowconfigure(0, weight=1)

        text_frame = tk.Frame(viewer_card, bg="#1F2937", highlightthickness=0)
        text_frame.grid(row=0, column=0, padx=18, pady=18, sticky="nsew")
        text_frame.grid_columnconfigure(0, weight=1)
        text_frame.grid_rowconfigure(0, weight=1)

        self.report_text = tk.Text(
            text_frame,
            wrap="none",
            bg="#0F172A",
            fg="#E2E8F0",
            insertbackground="#E2E8F0",
            relief="flat",
            font=("Consolas", 12),
            padx=18,
            pady=16
        )

        y_scroll = ttk.Scrollbar(text_frame, orient="vertical", command=self.report_text.yview)
        x_scroll = ttk.Scrollbar(text_frame, orient="horizontal", command=self.report_text.xview)

        self.report_text.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        self.report_text.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")

        self._set_report_text(
            "========== REPORTS ==========\n\n"
            "Choose one of the report options above to load the report here."
        )

    def _set_report_text(self, content):
        self.report_text.configure(state="normal")
        self.report_text.delete("1.0", tk.END)
        self.report_text.insert("1.0", content)
        self.report_text.configure(state="disabled")

    def _fetch_rows(self, query, params=None):
        conn = get_connection()
        cursor = conn.cursor()

        try:
            if params:
                cursor.execute(query, params)
            else:
                cursor.execute(query)
            return cursor.fetchall()
        finally:
            cursor.close()
            conn.close()

    def _show_report(self, title, content, row_count):
        self._set_report_text(content)
        self.summary_var.set(f"{title} loaded with {row_count} records")

    def _set_export_report(self, title, headers, rows):
        """Store the already displayed report data for reusable CSV/XLSX exporters."""
        self.current_report_title = title
        self.current_report_headers = tuple(headers)
        self.current_report_rows = [tuple(row) for row in rows]

    def _clear_export_report(self):
        self.current_report_title = None
        self.current_report_headers = ()
        self.current_report_rows = []

    def _require_report_for_export(self):
        if self.current_report_title:
            return True
        messagebox.showinfo("No Report Loaded", "Load a report before exporting it.")
        return False

    def _export_path(self, extension, label):
        safe_name = (self.current_report_title or "report").replace(" ", "_")
        return filedialog.asksaveasfilename(
            title=f"Export {self.current_report_title}",
            initialfile=f"{safe_name}.{extension}",
            defaultextension=f".{extension}",
            filetypes=[(label, f"*.{extension}")],
        )

    def export_csv(self):
        if not self._require_report_for_export():
            return
        path = self._export_path("csv", "CSV files")
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as file:
                writer = csv.writer(file)
                writer.writerow(self.current_report_headers)
                writer.writerows(self.current_report_rows)
            self.summary_var.set(f"{self.current_report_title} exported successfully.")
        except PermissionError:
            messagebox.showerror("Export Failed", "Permission was denied for the selected location.")
        except OSError as error:
            messagebox.showerror("Export Failed", f"The report could not be saved.\n\n{error}")
        except Exception as error:
            messagebox.showerror("Export Failed", f"Unexpected error while exporting the report.\n\n{error}")

    def export_excel(self):
        if not self._require_report_for_export():
            return
        path = self._export_path("xlsx", "Excel files")
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.worksheet.table import Table, TableStyleInfo

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = self.current_report_title[:31]
            worksheet.append(list(self.current_report_headers))
            for row in self.current_report_rows:
                worksheet.append(list(row))
            header_fill = PatternFill("solid", fgColor="1F2937")
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
            worksheet.freeze_panes = "A2"
            if self.current_report_rows:
                end_column = worksheet.cell(row=1, column=len(self.current_report_headers)).column_letter
                table = Table(displayName="ReportData", ref=f"A1:{end_column}{worksheet.max_row}")
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
                worksheet.add_table(table)
            for column_cells in worksheet.columns:
                width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 45)
                worksheet.column_dimensions[column_cells[0].column_letter].width = max(width, 12)
            workbook.save(path)
            self.summary_var.set(f"{self.current_report_title} exported successfully.")
        except PermissionError:
            messagebox.showerror("Export Failed", "Permission was denied for the selected location.")
        except ImportError:
            messagebox.showerror("Export Unavailable", "Excel export requires openpyxl. Install the project requirements and try again.")
        except OSError as error:
            messagebox.showerror("Export Failed", f"The report could not be saved.\n\n{error}")
        except Exception as error:
            messagebox.showerror("Export Failed", f"Unexpected error while exporting the report.\n\n{error}")

    def load_inventory_report(self):
        try:
            rows = self._fetch_rows(
                """
                SELECT
                    P.ProductName,
                    I.Quantity,
                    I.ReorderLevel
                FROM Inventory AS I
                INNER JOIN Products AS P
                    ON I.ProductID = P.ProductID
                ORDER BY P.ProductName
                """
            )

            lines = [
                "========== INVENTORY REPORT ==========",
                "",
                f"{'Product':<35}{'Quantity':<12}{'Reorder':<12}{'Status'}",
                "-" * 75,
            ]

            for row in rows:
                status = "LOW STOCK" if row.Quantity <= row.ReorderLevel else "IN STOCK"
                lines.append(
                    f"{str(row.ProductName):<35}"
                    f"{str(row.Quantity):<12}"
                    f"{str(row.ReorderLevel):<12}"
                    f"{status}"
                )

            self._set_export_report(
                "Inventory Report", ("Product", "Quantity", "Reorder Level", "Status"),
                [(row.ProductName, row.Quantity, row.ReorderLevel, "LOW STOCK" if row.Quantity <= row.ReorderLevel else "IN STOCK") for row in rows],
            )

            self._show_report("Inventory Report", "\n".join(lines), len(rows))
        except Exception as error:
            self._clear_export_report()
            self._set_report_text(f"Failed to load inventory report.\n\n{error}")
            self.summary_var.set("Failed to load inventory report")

    def load_sales_report(self):
        try:
            rows = self._fetch_rows(
                """
                SELECT
                    O.OrderID,
                    C.FirstName,
                    C.LastName,
                    O.OrderDate,
                    O.TotalAmount,
                    O.OrderStatus
                FROM Orders AS O
                INNER JOIN Customers AS C
                    ON O.CustomerID = C.CustomerID
                ORDER BY O.OrderID DESC
                """
            )

            lines = [
                "========== SALES REPORT ==========",
                "",
                f"{'Order ID':<10}{'Customer':<25}{'Date':<25}{'Amount':<15}{'Status'}",
                "-" * 90,
            ]

            for row in rows:
                customer = f"{row.FirstName} {row.LastName or ''}".strip()
                lines.append(
                    f"{str(row.OrderID):<10}"
                    f"{customer:<25}"
                    f"{str(row.OrderDate):<25}"
                    f"{str(row.TotalAmount):<15}"
                    f"{row.OrderStatus}"
                )

            self._set_export_report(
                "Sales Report", ("Order ID", "Customer", "Date", "Amount", "Status"),
                [(row.OrderID, f"{row.FirstName} {row.LastName or ''}".strip(), row.OrderDate, row.TotalAmount, row.OrderStatus) for row in rows],
            )

            self._show_report("Sales Report", "\n".join(lines), len(rows))
        except Exception as error:
            self._clear_export_report()
            self._set_report_text(f"Failed to load sales report.\n\n{error}")
            self.summary_var.set("Failed to load sales report")

    def load_product_report(self):
        try:
            rows = self._fetch_rows(
                """
                SELECT
                    P.ProductName,
                    C.CategoryName,
                    S.SupplierName,
                    P.SellingPrice
                FROM Products AS P
                INNER JOIN Categories AS C
                    ON P.CategoryID = C.CategoryID
                INNER JOIN Suppliers AS S
                    ON P.SupplierID = S.SupplierID
                ORDER BY P.ProductName
                """
            )

            lines = [
                "========== PRODUCT REPORT ==========",
                "",
                f"{'Product':<30}{'Category':<20}{'Supplier':<30}{'Price'}",
                "-" * 100,
            ]

            for row in rows:
                lines.append(
                    f"{str(row.ProductName):<30}"
                    f"{str(row.CategoryName):<20}"
                    f"{str(row.SupplierName):<30}"
                    f"{row.SellingPrice}"
                )

            self._set_export_report(
                "Product Report", ("Product", "Category", "Supplier", "Selling Price"),
                [(row.ProductName, row.CategoryName, row.SupplierName, row.SellingPrice) for row in rows],
            )

            self._show_report("Product Report", "\n".join(lines), len(rows))
        except Exception as error:
            self._clear_export_report()
            self._set_report_text(f"Failed to load product report.\n\n{error}")
            self.summary_var.set("Failed to load product report")

    def load_payment_report(self):
        try:
            rows = self._fetch_rows(
                """
                SELECT
                    P.PaymentID,
                    P.OrderID,
                    C.FirstName,
                    C.LastName,
                    P.PaymentMethod,
                    P.Amount,
                    P.PaymentStatus
                FROM Payments AS P
                INNER JOIN Orders AS O
                    ON P.OrderID = O.OrderID
                INNER JOIN Customers AS C
                    ON O.CustomerID = C.CustomerID
                ORDER BY P.PaymentID DESC
                """
            )

            lines = [
                "========== PAYMENT REPORT ==========",
                "",
                f"{'Payment':<10}{'Order':<10}{'Customer':<25}{'Method':<18}{'Amount':<15}{'Status'}",
                "-" * 100,
            ]

            for row in rows:
                customer = f"{row.FirstName} {row.LastName or ''}".strip()
                lines.append(
                    f"{str(row.PaymentID):<10}"
                    f"{str(row.OrderID):<10}"
                    f"{customer:<25}"
                    f"{str(row.PaymentMethod):<18}"
                    f"{str(row.Amount):<15}"
                    f"{row.PaymentStatus}"
                )

            self._set_export_report(
                "Payment Report", ("Payment ID", "Order ID", "Customer", "Method", "Amount", "Status"),
                [(row.PaymentID, row.OrderID, f"{row.FirstName} {row.LastName or ''}".strip(), row.PaymentMethod, row.Amount, row.PaymentStatus) for row in rows],
            )

            self._show_report("Payment Report", "\n".join(lines), len(rows))
        except Exception as error:
            self._clear_export_report()
            self._set_report_text(f"Failed to load payment report.\n\n{error}")
            self.summary_var.set("Failed to load payment report")

    def load_customer_report(self):
        try:
            rows = self._fetch_rows(
                """
                SELECT
                    C.CustomerID,
                    C.FirstName,
                    C.LastName,
                    COUNT(O.OrderID) AS TotalOrders,
                    ISNULL(SUM(O.TotalAmount), 0) AS TotalSpent
                FROM Customers AS C
                LEFT JOIN Orders AS O
                    ON C.CustomerID = O.CustomerID
                GROUP BY
                    C.CustomerID,
                    C.FirstName,
                    C.LastName
                ORDER BY TotalSpent DESC
                """
            )

            lines = [
                "========== CUSTOMER REPORT ==========",
                "",
                f"{'ID':<6}{'Customer':<30}{'Orders':<12}{'Total Spent'}",
                "-" * 70,
            ]

            for row in rows:
                customer = f"{row.FirstName} {row.LastName or ''}".strip()
                lines.append(
                    f"{str(row.CustomerID):<6}"
                    f"{customer:<30}"
                    f"{str(row.TotalOrders):<12}"
                    f"{row.TotalSpent}"
                )

            self._set_export_report(
                "Customer Report", ("Customer ID", "Customer", "Orders", "Total Spent"),
                [(row.CustomerID, f"{row.FirstName} {row.LastName or ''}".strip(), row.TotalOrders, row.TotalSpent) for row in rows],
            )

            self._show_report("Customer Report", "\n".join(lines), len(rows))
        except Exception as error:
            self._clear_export_report()
            self._set_report_text(f"Failed to load customer report.\n\n{error}")
            self.summary_var.set("Failed to load customer report")

    def load_supplier_report(self):
        try:
            rows = self._fetch_rows(
                """
                SELECT
                    S.SupplierID,
                    S.SupplierName,
                    COUNT(P.ProductID) AS ProductsSupplied
                FROM Suppliers AS S
                LEFT JOIN Products AS P
                    ON S.SupplierID = P.SupplierID
                GROUP BY
                    S.SupplierID,
                    S.SupplierName
                ORDER BY ProductsSupplied DESC
                """
            )

            lines = [
                "========== SUPPLIER REPORT ==========",
                "",
                f"{'ID':<6}{'Supplier':<40}{'Products Supplied'}",
                "-" * 65,
            ]

            for row in rows:
                lines.append(
                    f"{str(row.SupplierID):<6}"
                    f"{str(row.SupplierName):<40}"
                    f"{row.ProductsSupplied}"
                )

            self._set_export_report(
                "Supplier Report", ("Supplier ID", "Supplier", "Products Supplied"),
                [(row.SupplierID, row.SupplierName, row.ProductsSupplied) for row in rows],
            )

            self._show_report("Supplier Report", "\n".join(lines), len(rows))
        except Exception as error:
            self._clear_export_report()
            self._set_report_text(f"Failed to load supplier report.\n\n{error}")
            self.summary_var.set("Failed to load supplier report")
